import streamlit as st
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.styles import PatternFill, Font
import requests
import re
import csv
from pathlib import Path
from datetime import date, timedelta
from io import BytesIO, StringIO

# Edycja Nomenklatury Scalonej, względem której walidujemy kody CN.
# Lista w pliku cn_<rok>.txt (8-cyfrowe kody, jedna sztuka na linię).
# Źródło: GUS (stat.gov.pl), oficjalny plik CN dla Intrastatu.
# AKTUALIZACJA RAZ W ROKU: pobrać nowy plik CN, wygenerować cn_<rok>.txt, podbić rok.
CN_EDITION_YEAR = 2026

HIGHLIGHT_FILL = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
HIGHLIGHT_HEADER = PatternFill(start_color="4BACC6", end_color="4BACC6", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")

# Mapa ISO 3166-1: kod 3-literowy (alpha-3) -> kod 2-literowy (alpha-2)
ISO_A3_TO_A2 = {
    "AFG": "AF", "ALA": "AX", "ALB": "AL", "DZA": "DZ", "ASM": "AS", "AND": "AD",
    "AGO": "AO", "AIA": "AI", "ATA": "AQ", "ATG": "AG", "ARG": "AR", "ARM": "AM",
    "ABW": "AW", "AUS": "AU", "AUT": "AT", "AZE": "AZ", "BHS": "BS", "BHR": "BH",
    "BGD": "BD", "BRB": "BB", "BLR": "BY", "BEL": "BE", "BLZ": "BZ", "BEN": "BJ",
    "BMU": "BM", "BTN": "BT", "BOL": "BO", "BES": "BQ", "BIH": "BA", "BWA": "BW",
    "BVT": "BV", "BRA": "BR", "IOT": "IO", "BRN": "BN", "BGR": "BG", "BFA": "BF",
    "BDI": "BI", "CPV": "CV", "KHM": "KH", "CMR": "CM", "CAN": "CA", "CYM": "KY",
    "CAF": "CF", "TCD": "TD", "CHL": "CL", "CHN": "CN", "CXR": "CX", "CCK": "CC",
    "COL": "CO", "COM": "KM", "COG": "CG", "COD": "CD", "COK": "CK", "CRI": "CR",
    "CIV": "CI", "HRV": "HR", "CUB": "CU", "CUW": "CW", "CYP": "CY", "CZE": "CZ",
    "DNK": "DK", "DJI": "DJ", "DMA": "DM", "DOM": "DO", "ECU": "EC", "EGY": "EG",
    "SLV": "SV", "GNQ": "GQ", "ERI": "ER", "EST": "EE", "SWZ": "SZ", "ETH": "ET",
    "FLK": "FK", "FRO": "FO", "FJI": "FJ", "FIN": "FI", "FRA": "FR", "GUF": "GF",
    "PYF": "PF", "ATF": "TF", "GAB": "GA", "GMB": "GM", "GEO": "GE", "DEU": "DE",
    "GHA": "GH", "GIB": "GI", "GRC": "GR", "GRL": "GL", "GRD": "GD", "GLP": "GP",
    "GUM": "GU", "GTM": "GT", "GGY": "GG", "GIN": "GN", "GNB": "GW", "GUY": "GY",
    "HTI": "HT", "HMD": "HM", "VAT": "VA", "HND": "HN", "HKG": "HK", "HUN": "HU",
    "ISL": "IS", "IND": "IN", "IDN": "ID", "IRN": "IR", "IRQ": "IQ", "IRL": "IE",
    "IMN": "IM", "ISR": "IL", "ITA": "IT", "JAM": "JM", "JPN": "JP", "JEY": "JE",
    "JOR": "JO", "KAZ": "KZ", "KEN": "KE", "KIR": "KI", "PRK": "KP", "KOR": "KR",
    "KWT": "KW", "KGZ": "KG", "LAO": "LA", "LVA": "LV", "LBN": "LB", "LSO": "LS",
    "LBR": "LR", "LBY": "LY", "LIE": "LI", "LTU": "LT", "LUX": "LU", "MAC": "MO",
    "MDG": "MG", "MWI": "MW", "MYS": "MY", "MDV": "MV", "MLI": "ML", "MLT": "MT",
    "MHL": "MH", "MTQ": "MQ", "MRT": "MR", "MUS": "MU", "MYT": "YT", "MEX": "MX",
    "FSM": "FM", "MDA": "MD", "MCO": "MC", "MNG": "MN", "MNE": "ME", "MSR": "MS",
    "MAR": "MA", "MOZ": "MZ", "MMR": "MM", "NAM": "NA", "NRU": "NR", "NPL": "NP",
    "NLD": "NL", "NCL": "NC", "NZL": "NZ", "NIC": "NI", "NER": "NE", "NGA": "NG",
    "NIU": "NU", "NFK": "NF", "MKD": "MK", "MNP": "MP", "NOR": "NO", "OMN": "OM",
    "PAK": "PK", "PLW": "PW", "PSE": "PS", "PAN": "PA", "PNG": "PG", "PRY": "PY",
    "PER": "PE", "PHL": "PH", "PCN": "PN", "POL": "PL", "PRT": "PT", "PRI": "PR",
    "QAT": "QA", "REU": "RE", "ROU": "RO", "RUS": "RU", "RWA": "RW", "BLM": "BL",
    "SHN": "SH", "KNA": "KN", "LCA": "LC", "MAF": "MF", "SPM": "PM", "VCT": "VC",
    "WSM": "WS", "SMR": "SM", "STP": "ST", "SAU": "SA", "SEN": "SN", "SRB": "RS",
    "SYC": "SC", "SLE": "SL", "SGP": "SG", "SXM": "SX", "SVK": "SK", "SVN": "SI",
    "SLB": "SB", "SOM": "SO", "ZAF": "ZA", "SGS": "GS", "SSD": "SS", "ESP": "ES",
    "LKA": "LK", "SDN": "SD", "SUR": "SR", "SJM": "SJ", "SWE": "SE", "CHE": "CH",
    "SYR": "SY", "TWN": "TW", "TJK": "TJ", "TZA": "TZ", "THA": "TH", "TLS": "TL",
    "TGO": "TG", "TKL": "TK", "TON": "TO", "TTO": "TT", "TUN": "TN", "TUR": "TR",
    "TKM": "TM", "TCA": "TC", "TUV": "TV", "UGA": "UG", "UKR": "UA", "ARE": "AE",
    "GBR": "GB", "USA": "US", "UMI": "UM", "URY": "UY", "UZB": "UZ", "VUT": "VU",
    "VEN": "VE", "VNM": "VN", "VGB": "VG", "VIR": "VI", "WLF": "WF", "ESH": "EH",
    "YEM": "YE", "ZMB": "ZM", "ZWE": "ZW",
}


def normalize_country_code(val):
    """Konwertuje kod kraju alpha-3 (np. 'ITA') na alpha-2 (np. 'IT').
    Nieznane kody / wartości zwraca bez zmian."""
    if val is None:
        return None
    code = str(val).strip().upper()
    return ISO_A3_TO_A2.get(code, val)


# Aliasy nazw walut (poza kodami ISO z CURRENCIES) → kod ISO. Klucze WIELKIMI literami,
# bez spacji wiodących. Obsługa częstych form PL/EN spotykanych w plikach klientów.
CURRENCY_ALIASES = {
    "EURO": "EUR", "EUROS": "EUR", "€": "EUR",
    "DOLAR": "USD", "DOLAR AMERYKAŃSKI": "USD", "DOLAR USA": "USD",
    "DOLLAR": "USD", "US DOLLAR": "USD", "USD$": "USD", "$": "USD",
    "FUNT": "GBP", "FUNT SZTERLING": "GBP", "FUNT BRYTYJSKI": "GBP",
    "POUND": "GBP", "£": "GBP",
    "FRANK": "CHF", "FRANK SZWAJCARSKI": "CHF", "SWISS FRANC": "CHF",
    "KORONA CZESKA": "CZK", "KORONA DUŃSKA": "DKK", "KORONA NORWESKA": "NOK",
    "KORONA SZWEDZKA": "SEK",
    "FORINT": "HUF", "LEJ": "RON", "LEJ RUMUŃSKI": "RON",
    "HRYWNA": "UAH", "LIRA": "TRY", "LIRA TURECKA": "TRY",
    "JEN": "JPY", "YEN": "JPY", "YUAN": "CNY", "JUAN": "CNY", "RENMINBI": "CNY",
    "ZŁOTY": "PLN", "ZLOTY": "PLN", "ZŁ": "PLN", "ZL": "PLN", "PLN": "PLN",
}


def normalize_currency(val):
    """Sprowadza wartość komórki do kodu waluty ISO (np. 'eur' / 'euro' → 'EUR').
    PLN rozpoznaje (kurs 1, brak zapytania do API). Puste / nieznane → None."""
    if val is None:
        return None
    code = str(val).strip().upper()
    if code == "":
        return None
    if code in CURRENCIES or code == "PLN":
        return code
    return CURRENCY_ALIASES.get(code)


def first_line_only(val):
    """Jeśli komórka zawiera złamanie linii (Alt+Enter), zwraca tylko pierwszą linię.
    Pozostałe wartości zwraca bez zmian."""
    if isinstance(val, str) and ("\n" in val or "\r" in val):
        lines = val.splitlines()
        return lines[0] if lines else ""
    return val


def _coerce_csv_value(s: str):
    """Zamienia tekst z CSV na liczbę, jeśli to czysta liczba. Zachowuje wiodące zera
    jako tekst (kody CN/ID), obsługuje przecinek dziesiętny (1234,56)."""
    s = s.strip()
    if s == "":
        return None
    if re.fullmatch(r"-?\d+", s):
        body = s.lstrip("-")
        if len(body) > 1 and body.startswith("0"):
            return s  # wiodące zero = prawdopodobnie kod, zostaw tekstem
        return int(s)
    t = s.replace(" ", "")
    if re.fullmatch(r"-?\d+[.,]\d+", t):
        return float(t.replace(",", "."))
    return s


def read_csv_to_workbook(raw: bytes):
    """Wczytuje CSV (bytes) do skoroszytu openpyxl w pamięci.
    Zwraca (wb, delimiter). Auto-wykrywa kodowanie i separator."""
    text = None
    for enc in ("utf-8-sig", "cp1250", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = raw.decode("utf-8", errors="replace")

    sample = text[:5000]
    counts = {d: sample.count(d) for d in (";", "\t", ",")}
    delimiter = max(counts, key=counts.get) if max(counts.values()) > 0 else ","

    wb = openpyxl.Workbook()
    ws = wb.active
    reader = csv.reader(StringIO(text), delimiter=delimiter)
    for ri, row in enumerate(reader, start=1):
        for ci, val in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=_coerce_csv_value(val))
    return wb, delimiter


def workbook_to_csv_bytes(ws, delimiter: str) -> bytes:
    """Eksportuje arkusz do CSV (bytes, UTF-8 z BOM). Liczby dziesiętne: przecinek
    gdy separator to ';', kropka gdy ','."""
    dec = "," if delimiter == ";" else "."
    buf = StringIO()
    writer = csv.writer(buf, delimiter=delimiter, lineterminator="\r\n")
    for row in ws.iter_rows(values_only=True):
        out = []
        for v in row:
            if v is None:
                out.append("")
            elif isinstance(v, float):
                out.append(f"{v:.6f}".rstrip("0").rstrip(".").replace(".", dec))
            else:
                out.append(str(v))
        writer.writerow(out)
    return buf.getvalue().encode("utf-8-sig")


def load_cn_codes(year: int) -> frozenset:
    """Wczytuje zbiór obowiązujących 8-cyfrowych kodów CN z pliku cn_<rok>.txt.
    Zwraca pusty zbiór, jeśli pliku brak (walidacja wtedy nieaktywna)."""
    path = Path(__file__).with_name(f"cn_{year}.txt")
    if not path.exists():
        return frozenset()
    return frozenset(
        line.strip() for line in path.read_text(encoding="utf-8").splitlines() if line.strip()
    )


def normalize_cn(val) -> str:
    """Kanoniczna postać kodu CN: 8 cyfr (CN) lub 10 cyfr (TARIC), jeśli się da.
    Excel gubi wiodące zero przy kodach z rozdziałów 01-09 (01012100 -> liczba 1012100,
    a 0102901000 -> 102901000) — 7-cyfrowy uzupełniamy do 8, 9-cyfrowy do 10.
    Liczbę zmiennoprzecinkową bez części ułamkowej (np. 84713000.0) sprowadzamy do int,
    by „.0" nie zafałszowało liczby cyfr."""
    if isinstance(val, float) and val.is_integer():
        val = int(val)
    digits = re.sub(r"\D", "", str(val)) if val is not None else ""
    if len(digits) == 7:
        digits = digits.zfill(8)
    elif len(digits) == 9:
        digits = digits.zfill(10)
    return digits


def cn_status(code: str, valid_cn: frozenset):
    """Status kanonicznego kodu. Możliwe zwroty:
      None                                   - pusty
      'OK'                                   - 8 cyfr (CN), w obowiązującej edycji
      'OK (10-cyfr → sprawdzono prefiks 8)'  - 10 cyfr (TARIC), prefiks 8 cyfr w edycji
      'nieaktualny'                          - 8 cyfr, poza edycją
      'nieaktualny (10-cyfr → prefiks 8 poza edycją)' - 10 cyfr, prefiks poza edycją
      'błędny format'                        - inna długość niż 8/10 cyfr
    Dla kodu 10-cyfrowego walidujemy prefiks 8 cyfr (= kod CN wewnątrz kodu TARIC)."""
    if not code:
        return None
    if len(code) == 8:
        return "OK" if code in valid_cn else "nieaktualny"
    if len(code) == 10:
        if code[:8] in valid_cn:
            return "OK (10-cyfr → sprawdzono prefiks 8)"
        return "nieaktualny (10-cyfr → prefiks 8 poza edycją)"
    return "błędny format"


def cn_is_valid(status) -> bool:
    """Czy status oznacza kod NIE wymagający poprawy (pusty lub poprawny — też 10-cyfrowy).
    Kontrakt: każdy status zaczynający się od 'OK' = poprawny (świadome sprzężenie
    z etykietą; przy zmianie nazwy statusu OK trzeba zachować ten prefiks)."""
    return status is None or status.startswith("OK")


def cn_badge(status) -> str:
    """Etykieta statusu z ikoną do UI (odporna na dowolny tekst statusu, też 10-cyfrowy)."""
    if status == "poprawiony" or cn_is_valid(status):
        return f"✅ {status}"
    return f"❌ {status}"


def cn_outcome(original, replacement, valid_cn: frozenset):
    """JEDYNE źródło prawdy dla decyzji o kodzie CN — używane i w UI, i przy zapisie pliku.
    Zwraca (status, wartość_do_zapisania_w_komórce_CN):
    - poprawny/pusty oryginał      -> (OK/None, oryginał)
    - niepoprawny bez zamiennika   -> (status oryginału, oryginał)
    - niepoprawny + zamiennik OK    -> ('poprawiony', zamiennik)
    - niepoprawny + zamiennik zły   -> (status zamiennika, ORYGINAŁ — złego zamiennika NIE zapisujemy)"""
    status = cn_status(normalize_cn(original), valid_cn)
    if cn_is_valid(status):
        return status, original
    if replacement is None or str(replacement).strip() == "":
        return status, original
    repl = str(replacement).strip()
    repl_status = cn_status(normalize_cn(repl), valid_cn)
    # zamiennik akceptujemy tylko jeśli sam jest poprawnym kodem (8 lub 10 cyfr);
    # None = zamiennik bez cyfr/pusty → odrzucamy (cn_is_valid(None) jest True, stąd jawny warunek)
    if repl_status is not None and cn_is_valid(repl_status):
        return "poprawiony", repl
    # zły zamiennik: zostaje oryginał i jego status (nie zapisujemy złego kodu)
    return status, original


def check_cn_code(val, valid_cn: frozenset):
    """Status pojedynczego kodu bez zamiennika (zachowane dla zgodności/testów)."""
    return cn_outcome(val, None, valid_cn)[0]


def extract_vat_country(val):
    """Wyciąga 2-literowy kod kraju z numeru VAT (np. 'DE123456789' -> 'DE').
    EL (prefiks VAT Grecji) zamienia na GR. Pozostałe prefiksy zgodne z ISO 3166.
    Brak czytelnego prefiksu (np. sam numer) -> None (pusta komórka)."""
    if val is None:
        return None
    prefix = str(val).strip().upper()[:2]
    if len(prefix) == 2 and prefix.isalpha():
        return "GR" if prefix == "EL" else prefix
    return None


def fetch_nbp_rates(date_from: date, date_to: date, currency: str = "eur") -> dict[date, float]:
    """Pobiera wszystkie kursy waluty/PLN z NBP w podanym zakresie dat (jedno zapytanie)."""
    url = f"https://api.nbp.pl/api/exchangerates/rates/a/{currency.lower()}/{date_from}/{date_to}/?format=json"
    try:
        resp = requests.get(url, timeout=30)
        if resp.status_code == 200:
            data = resp.json()
            return {
                date.fromisoformat(r["effectiveDate"]): r["mid"]
                for r in data["rates"]
            }
    except requests.RequestException:
        pass
    return {}


def fetch_ecb_rates(date_from: date, date_to: date) -> dict[date, float]:
    """Pobiera wszystkie kursy EUR/PLN z ECB w podanym zakresie dat (jedno zapytanie)."""
    url = (
        f"https://data-api.ecb.europa.eu/service/data/EXR/"
        f"D.PLN.EUR.SP00.A?startPeriod={date_from}&endPeriod={date_to}"
        f"&format=csvdata"
    )
    rates = {}
    try:
        resp = requests.get(url, timeout=30)
        if resp.status_code == 200 and "OBS_VALUE" in resp.text:
            lines = resp.text.strip().split("\n")
            if len(lines) >= 2:
                header = lines[0].split(",")
                obs_idx = header.index("OBS_VALUE")
                date_idx = header.index("TIME_PERIOD")
                for line in lines[1:]:
                    values = line.split(",")
                    try:
                        rates[date.fromisoformat(values[date_idx])] = float(values[obs_idx])
                    except (ValueError, IndexError):
                        continue
    except requests.RequestException:
        pass
    return rates


def find_previous_rate(target_date: date, all_rates: dict[date, float]) -> tuple[float | None, date | None]:
    """Znajduje kurs z ostatniego dnia roboczego PRZED podaną datą."""
    check_date = target_date - timedelta(days=1)
    for _ in range(10):
        if check_date in all_rates:
            return all_rates[check_date], check_date
        check_date -= timedelta(days=1)
    return None, None


def penultimate_wednesday(year: int, month: int) -> date:
    """Przedostatnia środa danego miesiąca (NBP publikuje wtedy kurs celny)."""
    import calendar
    days = calendar.monthrange(year, month)[1]
    weds = [date(year, month, d) for d in range(1, days + 1) if date(year, month, d).weekday() == 2]
    return weds[-2]


def customs_reference_date(d: date) -> date:
    """Data publikacji kursu celnego obowiązującego dla MIESIĄCA daty `d`:
    przedostatnia środa POPRZEDNIEGO miesiąca kalendarzowego.
    (Kurs ogłoszony w przedostatnią środę miesiąca M obowiązuje przez cały M+1.)"""
    last_prev = d.replace(day=1) - timedelta(days=1)  # ostatni dzień poprzedniego miesiąca
    return penultimate_wednesday(last_prev.year, last_prev.month)


def rate_on_or_before(target: date, all_rates: dict[date, float]) -> float | None:
    """Kurs z dnia `target`, a jeśli tego dnia brak notowania (święto) — z ostatniego
    dnia roboczego PRZED nim. Używane dla kursu celnego (data = przedostatnia środa)."""
    check = target
    for _ in range(10):
        if check in all_rates:
            return all_rates[check]
        check -= timedelta(days=1)
    return None


def parse_date_value(val) -> date | None:
    """Próbuje sparsować wartość komórki jako datę."""
    from datetime import datetime
    if hasattr(val, "date"):  # datetime
        return val.date()
    if isinstance(val, date):
        return val
    # YYYYMMDD jako liczba (np. 20260417)
    if isinstance(val, int):
        s = str(val)
        if len(s) == 8:
            try:
                return datetime.strptime(s, "%Y%m%d").date()
            except ValueError:
                return None
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y", "%Y%m%d"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
    return None


def update_table_refs(ws, insert_col: int, col_name: str = ""):
    """Aktualizuje zakresy i kolumny tabel strukturalnych po wstawieniu kolumny."""
    from openpyxl.worksheet.table import TableColumn
    for table in ws.tables.values():
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        if insert_col > max_col:
            continue
        # Pozycja wstawienia względem tabeli (0-based)
        insert_pos = insert_col - min_col
        if insert_col < min_col:
            min_col += 1
            # Kolumna wstawiona przed tabelą — nie dodajemy kolumny do definicji
        else:
            # Kolumna wstawiona wewnątrz tabeli — dodaj do tableColumns
            new_tc = TableColumn(id=max_col - min_col + 2, name=col_name or f"Column{max_col + 1}")
            cols = list(table.tableColumns)
            cols.insert(insert_pos, new_tc)
            # Przenumeruj ID kolumn
            for i, tc in enumerate(cols):
                tc.id = i + 1
            table.tableColumns = cols
        max_col += 1
        table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def process_workbook(wb, sheet_name: str, col_idx: int, source: str, currency: str, amount_col_idx: int | None, progress_bar, country_col_idx: int | None = None, trim_multiline: bool = False, vat_col_idx: int | None = None, cn_col_idx: int | None = None, valid_cn: frozenset = frozenset(), cn_replacements: dict | None = None, currency_col_idx: int | None = None, rate_basis: str = "daily"):
    """Przetwarza arkusz — wstawia kolumny z kursami obok kolumny dat."""
    ws = wb[sheet_name]
    cn_replacements = cn_replacements or {}

    # Utnij wielolinijkowy tekst do pierwszej linii (przed wstawianiem kolumn).
    # Zmienia tylko komórki ze złamaniem linii (Alt+Enter), reszta bez zmian.
    if trim_multiline:
        for row in ws.iter_rows():
            for cell in row:
                cell.value = first_line_only(cell.value)

    # Kolumny pochodne (kod kraju ISO-2, kod kraju z VAT) wstawiamy PRZED kolumnami
    # z kursami. Wspólny helper koryguje WSZYSTKIE śledzone indeksy po każdym
    # wstawieniu — to eliminuje ciche błędy przesunięcia kolumn (4 wstawienia).
    idx = {"date": col_idx, "amount": amount_col_idx, "country": country_col_idx, "vat": vat_col_idx, "cn": cn_col_idx, "currency": currency_col_idx}

    def insert_derived_column(src_key, transform, header_suffix, default_header):
        src = idx[src_key]
        if src is None:
            return
        src_header = ws.cell(row=1, column=src).value
        values = {r: ws.cell(row=r, column=src).value for r in range(2, ws.max_row + 1)}
        out_col = src + 1
        ws.insert_cols(out_col)
        out_header = f"{src_header} {header_suffix}" if src_header else default_header
        update_table_refs(ws, out_col, out_header)
        hc = ws.cell(row=1, column=out_col, value=out_header)
        hc.fill = HIGHLIGHT_HEADER
        hc.font = HEADER_FONT
        for r in range(2, ws.max_row + 1):
            oc = ws.cell(row=r, column=out_col, value=transform(values.get(r)))
            oc.fill = HIGHLIGHT_FILL
        # Korekta wszystkich indeksów na/za pozycją wstawienia
        for k in idx:
            if idx[k] is not None and idx[k] >= out_col:
                idx[k] += 1

    insert_derived_column("country", normalize_country_code, "(ISO-2)", "Kod kraju (ISO-2)")
    insert_derived_column("vat", extract_vat_country, "(kod kraju)", "Kod kraju z VAT")

    # CN: walidacja + ewentualna podmiana kodów. Specjalny blok (nie generyczny helper),
    # bo poza nową kolumną ze statusem podmienia też kod w kolumnie ŹRÓDŁOWEJ.
    if idx["cn"] is not None and valid_cn:
        src = idx["cn"]
        src_header = ws.cell(row=1, column=src).value
        values = {r: ws.cell(row=r, column=src).value for r in range(2, ws.max_row + 1)}
        out_col = src + 1
        ws.insert_cols(out_col)
        out_header = f"{src_header} (status CN {CN_EDITION_YEAR})" if src_header else f"Status CN {CN_EDITION_YEAR}"
        update_table_refs(ws, out_col, out_header)
        hc = ws.cell(row=1, column=out_col, value=out_header)
        hc.fill = HIGHLIGHT_HEADER
        hc.font = HEADER_FONT
        for r in range(2, ws.max_row + 1):
            original = values.get(r)
            status, cell_value = cn_outcome(original, cn_replacements.get(normalize_cn(original)), valid_cn)
            if cell_value != original:
                ws.cell(row=r, column=src, value=cell_value)
            sc = ws.cell(row=r, column=out_col, value=status)
            sc.fill = HIGHLIGHT_FILL
        for k in idx:
            if idx[k] is not None and idx[k] >= out_col:
                idx[k] += 1

    col_idx = idx["date"]
    amount_col_idx = idx["amount"]
    currency_col_idx = idx["currency"]
    per_row_currency = currency_col_idx is not None

    customs = rate_basis == "monthly_customs"

    # Wstaw kolumnę z kursem zaraz po kolumnie z datami
    rate_col = col_idx + 1
    if customs:
        source_label = "kurs celny NBP"  # kurs celny zawsze z NBP (ECB nie publikuje kursu celnego)
    else:
        source_label = "kurs NBP" if source == "NBP" else "kurs EBC"
    # W trybie per-wiersz waluta jest różna w wierszach — nagłówek bez kodu waluty.
    rate_col_name = f"{source_label}/PLN" if per_row_currency else f"{source_label} {currency}/PLN"
    ws.insert_cols(rate_col)
    update_table_refs(ws, rate_col, rate_col_name)
    header_cell = ws.cell(row=1, column=rate_col, value=rate_col_name)
    header_cell.fill = HIGHLIGHT_HEADER
    header_cell.font = HEADER_FONT

    # Korekta indeksów kolumn położonych za wstawioną kolumną kursu
    if amount_col_idx is not None and amount_col_idx >= rate_col:
        amount_col_idx += 1
    if currency_col_idx is not None and currency_col_idx >= rate_col:
        currency_col_idx += 1

    new_cols = [rate_col]

    pln_col = None
    if amount_col_idx is not None:
        pln_col = rate_col + 1
        ws.insert_cols(pln_col)
        update_table_refs(ws, pln_col, "PLN przeliczone")
        pln_header = ws.cell(row=1, column=pln_col, value="PLN przeliczone")
        pln_header.fill = HIGHLIGHT_HEADER
        pln_header.font = HEADER_FONT
        new_cols.append(pln_col)
        # Skoryguj ponownie indeksy jeśli są za drugą wstawioną kolumną
        if amount_col_idx >= pln_col:
            amount_col_idx += 1
        if currency_col_idx is not None and currency_col_idx >= pln_col:
            currency_col_idx += 1

    total_rows = ws.max_row - 1
    if total_rows <= 0:
        return wb

    # Zbierz wszystkie daty z arkusza
    all_dates = []
    for row_num in range(2, ws.max_row + 1):
        parsed = parse_date_value(ws.cell(row=row_num, column=col_idx).value)
        if parsed:
            all_dates.append(parsed)

    progress_bar.progress(0.05, "Pobieram kursy z API...")

    # Pobierz kursy hurtowo dla jednego zakresu dat.
    # Tryb per-wiersz: jedno zapytanie na KAŻDĄ napotkaną walutę (PLN pomijamy — kurs 1).
    # Tryb jednej waluty: jedno zapytanie dla wybranej waluty.
    all_rates = {}                 # tryb jednej waluty
    rates_by_currency = {}         # tryb per-wiersz: kod waluty -> {data: kurs}
    # ECB tylko dla kursu dziennego EUR; kurs celny zawsze z NBP
    use_ecb = (not customs) and source == "ECB"
    if all_dates:
        if customs:
            # zakres musi sięgać najwcześniejszej przedostatniej środy (kurs celny)
            date_from = min(customs_reference_date(d) for d in all_dates) - timedelta(days=10)
        else:
            date_from = min(all_dates) - timedelta(days=15)
        date_to = max(all_dates)
        if per_row_currency:
            currencies_needed = set()
            for row_num in range(2, ws.max_row + 1):
                cur = normalize_currency(ws.cell(row=row_num, column=currency_col_idx).value)
                if cur and cur != "PLN":
                    currencies_needed.add(cur)
            for cur in currencies_needed:
                if cur == "EUR" and use_ecb:
                    rates_by_currency[cur] = fetch_ecb_rates(date_from, date_to)
                else:
                    rates_by_currency[cur] = fetch_nbp_rates(date_from, date_to, cur)
        elif use_ecb:
            all_rates = fetch_ecb_rates(date_from, date_to)
        else:
            all_rates = fetch_nbp_rates(date_from, date_to, currency)

    progress_bar.progress(0.3, "Wstawiam kursy do arkusza...")

    # Wstaw kursy do arkusza
    for i, row_num in enumerate(range(2, ws.max_row + 1)):
        parsed_date = parse_date_value(ws.cell(row=row_num, column=col_idx).value)

        # Ustal walutę i kurs dla wiersza
        row_cur = normalize_currency(ws.cell(row=row_num, column=currency_col_idx).value) if per_row_currency else currency
        if per_row_currency and row_cur is None:
            rate, rate_msg = None, "Nieznana waluta"
        elif row_cur == "PLN":
            rate, rate_msg = 1, None  # PLN/PLN = 1, niezależnie od daty
        elif parsed_date:
            row_rates = rates_by_currency.get(row_cur, {}) if per_row_currency else all_rates
            if customs:
                # kurs celny: ten sam dla całego miesiąca transakcji (przedostatnia środa M-1)
                rate = rate_on_or_before(customs_reference_date(parsed_date), row_rates)
            else:
                rate, _ = find_previous_rate(parsed_date, row_rates)
            rate_msg = None if rate is not None else "Brak kursu"
        else:
            rate, rate_msg = None, "Błędna data"

        if rate is not None:
            ws.cell(row=row_num, column=rate_col, value=rate)
            # Przelicz kwotę na PLN
            if pln_col and amount_col_idx:
                amount = ws.cell(row=row_num, column=amount_col_idx).value
                if isinstance(amount, (int, float)):
                    pln_cell = ws.cell(row=row_num, column=pln_col, value=round(amount * rate))
                    pln_cell.number_format = "0"
                else:
                    ws.cell(row=row_num, column=pln_col, value="Brak kwoty")
        else:
            ws.cell(row=row_num, column=rate_col, value=rate_msg)
            if pln_col:
                ws.cell(row=row_num, column=pln_col, value=rate_msg)

        # Highlight nowych kolumn
        for c in new_cols:
            ws.cell(row=row_num, column=c).fill = HIGHLIGHT_FILL

        progress_bar.progress(0.3 + 0.7 * (i + 1) / total_rows)

    return wb


# ---- STREAMLIT UI ----

st.set_page_config(page_title="FX Tool — Kursy walut", page_icon="💱", layout="wide")

CURRENCIES = {
    "EUR": "euro", "USD": "dolar amerykański", "GBP": "funt szterling",
    "CHF": "frank szwajcarski", "CZK": "korona czeska", "DKK": "korona duńska",
    "NOK": "korona norweska", "SEK": "korona szwedzka", "HUF": "forint (Węgry)",
    "RON": "lej rumuński", "UAH": "hrywna (Ukraina)", "TRY": "lira turecka",
    "CAD": "dolar kanadyjski", "AUD": "dolar australijski", "JPY": "jen (Japonia)",
    "CNY": "yuan renminbi (Chiny)", "HKD": "dolar Hongkongu", "SGD": "dolar singapurski",
    "NZD": "dolar nowozelandzki", "MXN": "peso meksykańskie", "BRL": "real (Brazylia)",
    "ZAR": "rand (RPA)", "KRW": "won (Korea Płd.)", "INR": "rupia indyjska",
    "ILS": "nowy izraelski szekel", "THB": "bat (Tajlandia)", "PHP": "peso filipińskie",
    "IDR": "rupia indonezyjska", "MYR": "ringgit (Malezja)", "CLP": "peso chilijskie",
    "ISK": "korona islandzka", "XDR": "SDR (MFW)",
}

@st.cache_data
def get_valid_cn(year: int) -> frozenset:
    return load_cn_codes(year)


VALID_CN = get_valid_cn(CN_EDITION_YEAR)

st.title("FX Tool — Kursy walut do Excel")
st.markdown("Załaduj plik Excel lub CSV, wskaż kolumnę z datami, a aplikacja wstawi kurs wybranej waluty/PLN z dnia poprzedzającego.")

st.warning("Plik wyjściowy zawiera wartości zamiast formuł. Oryginalny plik nie jest modyfikowany.")

# Upload pliku
uploaded_file = st.file_uploader(
    "Wybierz plik Excel lub CSV", type=["xlsx", "csv"],
    help="Obsługiwane formaty: .xlsx (Excel) i .csv. Dla CSV separator (`;` / tab / `,`) i kodowanie są wykrywane automatycznie, a plik wynikowy też powstaje w formacie CSV. Oryginalny plik nigdy nie jest modyfikowany — pobierasz nową kopię.",
)

if uploaded_file is not None:
    is_csv = uploaded_file.name.lower().endswith(".csv")
    csv_delimiter = ","
    if is_csv:
        wb, csv_delimiter = read_csv_to_workbook(uploaded_file.getvalue())
        st.caption(f"Wczytano CSV (separator: „{csv_delimiter}”). Plik wynikowy też będzie w formacie CSV.")
    else:
        file_bytes = BytesIO(uploaded_file.getvalue())  # getvalue() odporne na wielokrotne przeładowania
        wb = openpyxl.load_workbook(file_bytes, data_only=True, rich_text=False)

    # Wybór arkusza
    sheet_names = wb.sheetnames
    if len(sheet_names) == 1:
        sheet_name = sheet_names[0]
    else:
        sheet_name = st.selectbox(
            "Wybierz arkusz", sheet_names,
            help="Arkusz, którego dane zostaną przetworzone. Pozostałe arkusze w pliku nie są ruszane.",
        )

    ws = wb[sheet_name]

    # Podgląd danych
    st.subheader("Podgląd danych")
    preview_data = []
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        headers.append(str(val) if val else f"Kolumna {get_column_letter(col)}")

    for row in range(2, min(ws.max_row + 1, 12)):  # max 10 wierszy podglądu
        row_data = {}
        for col in range(1, ws.max_column + 1):
            row_data[headers[col - 1]] = ws.cell(row=row, column=col).value
        preview_data.append(row_data)

    st.dataframe(preview_data, use_container_width=True)

    # Wybór parametrów
    col1, col2 = st.columns(2)

    with col1:
        date_column = st.selectbox(
            "Wskaż kolumnę z datami", headers,
            help="Data transakcji/faktury — na jej podstawie dobierany jest kurs (przy kursie dziennym z dnia poprzedzającego, przy kursie celnym z miesiąca tej daty). Kurs trafia do nowej kolumny wstawionej zaraz obok.",
        )
        col_idx = headers.index(date_column) + 1

    with col2:
        currency_mode = st.radio(
            "Sposób doboru waluty",
            ["Jedna waluta dla całego pliku", "Waluta z kolumny (per wiersz)"],
            help="**Jedna waluta** — wszystkie kwoty w pliku są w tej samej walucie (wybierasz ją niżej). **Waluta z kolumny** — każdy wiersz może mieć inną walutę; wskazujesz kolumnę, z której odczytywany jest kod waluty (np. EUR, USD, PLN). PLN przelicza się kursem 1.",
        )
    per_row_currency = currency_mode.startswith("Waluta z kolumny")

    rate_basis_choice = st.radio(
        "Podstawa kursu",
        ["Kurs z dnia poprzedzającego (zasady VAT)", "Kurs celny — stały na miesiąc (zasady celne)"],
        horizontal=True,
        help=(
            "**Zasady VAT** — kurs średni NBP (lub EBC dla EUR) z ostatniego dnia roboczego "
            "**przed** datą każdej transakcji. **Zasady celne** — kurs celny NBP ogłaszany w "
            "**przedostatnią środę** miesiąca i obowiązujący przez **cały następny miesiąc** "
            "kalendarzowy; aplikacja stosuje go do wszystkich transakcji z danego miesiąca. "
            "Kurs celny zawsze z NBP (EBC nie publikuje kursu celnego). "
            "Wybierz jedną metodę konsekwentnie — nie miesza się ich per faktura."
        ),
    )
    customs = rate_basis_choice.startswith("Kurs celny")
    rate_basis = "monthly_customs" if customs else "daily"

    col3, col4 = st.columns(2)

    with col3:
        if per_row_currency:
            currency_col_options = headers
            # Auto-wykryj kolumnę z walutą (nazwa zawiera "waluta"/"currency"/"ccy")
            currency_col_default = 0
            for i, h in enumerate(headers):
                hl = h.strip().lower()
                if "waluta" in hl or "currency" in hl or hl in ("ccy", "cur", "kod waluty"):
                    currency_col_default = i
                    break
            currency_column = st.selectbox(
                "Kolumna z walutą", currency_col_options, index=currency_col_default,
                help="Kolumna zawierająca kod/nazwę waluty dla każdego wiersza (EUR, USD, eur, euro, PLN...). Dla każdego wiersza kurs dobierany jest wg tej waluty. Nierozpoznane wartości → „Nieznana waluta”. PLN → kurs 1.",
            )
            currency_col_idx = headers.index(currency_column) + 1
            currency = "EUR"  # nieużywane w trybie per-wiersz (placeholder)
        else:
            currency_options = [f"{code} — {name}" for code, name in CURRENCIES.items()]
            currency_choice = st.selectbox(
                "Waluta", currency_options,
                help="Waluta, której kurs do PLN zostanie pobrany. Dla **EUR** możesz wybrać źródło NBP lub EBC; pozostałe waluty zawsze z NBP.",
            )
            currency = currency_choice.split(" — ")[0]
            currency_col_idx = None

    with col4:
        if customs:
            source = "NBP"
            st.info("Źródło: **NBP** (kurs celny — EBC nie dotyczy)")
        elif per_row_currency:
            source = st.radio(
                "Źródło kursu (dotyczy EUR)", ["NBP", "ECB"], horizontal=True,
                help="W trybie per-wiersz wybór źródła dotyczy **tylko wierszy w EUR**; pozostałe waluty zawsze z NBP. **NBP** — tabela A. **ECB/EBC** — kursy referencyjne Europejskiego Banku Centralnego.",
            )
        elif currency == "EUR":
            source = st.radio(
                "Źródło kursu", ["NBP", "ECB"], horizontal=True,
                help="**NBP** — tabela A Narodowego Banku Polskiego. **ECB/EBC** — kursy referencyjne Europejskiego Banku Centralnego. Wybór dotyczy wyłącznie EUR.",
            )
        else:
            source = "NBP"
            st.info("Źródło: **NBP** (EBC dostępne tylko dla EUR)")

    amount_label = "waluta z kolumny" if per_row_currency else currency
    amount_options = ["— nie przeliczaj —"] + headers
    amount_column = st.selectbox(
        f"Kolumna z kwotami ({amount_label})", amount_options,
        help=f"Kwoty ({amount_label}) z tej kolumny zostaną przeliczone na PLN po pobranym kursie (zaokrąglone do pełnych złotych) i wpisane w nowej kolumnie obok. „— nie przeliczaj —” = tylko wstaw kurs, bez przeliczania kwot.",
    )
    amount_col_idx = headers.index(amount_column) + 1 if amount_column != "— nie przeliczaj —" else None

    # ---- Pre-skan wykrytych walut (tylko tryb per-wiersz) ----
    if per_row_currency:
        ws_cur = wb[sheet_name]
        found_cur = {}      # rozpoznany kod -> liczność
        unknown_cur = {}    # surowa wartość -> liczność
        for r in range(2, ws_cur.max_row + 1):
            raw = ws_cur.cell(row=r, column=currency_col_idx).value
            if raw is None or str(raw).strip() == "":
                continue
            code = normalize_currency(raw)
            if code:
                found_cur[code] = found_cur.get(code, 0) + 1
            else:
                key = str(raw).strip()
                unknown_cur[key] = unknown_cur.get(key, 0) + 1
        with st.container(border=True):
            if found_cur:
                summary = " · ".join(f"{c} ({n}×)" for c, n in sorted(found_cur.items()))
                st.markdown(f"**Wykryte waluty w kolumnie „{currency_column}”:** {summary}")
            if unknown_cur:
                u = " · ".join(f"„{k}” ({n}×)" for k, n in sorted(unknown_cur.items()))
                st.warning(f"⚠️ Nierozpoznane wartości: {u} — te wiersze dostaną status „Nieznana waluta” (bez przeliczenia).")
            elif found_cur:
                st.success("Wszystkie wartości w kolumnie waluty zostały rozpoznane.")
            else:
                st.info("Kolumna waluty jest pusta.")

    # ---- Sekcja specyficzna dla klienta: Fluiconnecto ----
    st.markdown("")
    with st.container(border=True):
        st.markdown("### :orange[🔶 Obróbka dla Fluiconnecto]")
        st.caption(
            "Operacje specyficzne dla klienta **Fluiconnecto**. "
            "Włączaj tylko przy plikach tego klienta."
        )

        fc1, fc2 = st.columns(2)
        with fc1:
            country_options = ["— nie normalizuj —"] + headers
            # Auto-wykryj kolumnę OrigCountryRegionId (bez względu na wielkość liter)
            country_default = 0
            for i, h in enumerate(headers):
                if h.strip().lower() == "origcountryregionid":
                    country_default = i + 1
                    break
            country_column = st.selectbox(
                "Kolumna z kodem kraju (alpha-3 → alpha-2)", country_options, index=country_default,
                help="Kody krajów 3-literowe (ISO alpha-3) zostaną zamienione na 2-literowe (alpha-2, np. POL → PL) w nowej kolumnie obok. Auto-wykrywa kolumnę „OrigCountryRegionId”.",
            )
            country_col_idx = headers.index(country_column) + 1 if country_column != "— nie normalizuj —" else None

        with fc2:
            vat_options = ["— nie wyciągaj —"] + headers
            # Auto-wykryj kolumnę z VAT (nazwa zawiera "vat")
            vat_default = 0
            for i, h in enumerate(headers):
                if "vat" in h.strip().lower():
                    vat_default = i + 1
                    break
            vat_column = st.selectbox(
                "Kolumna z numerem VAT → kod kraju", vat_options, index=vat_default,
                help="Z numeru VAT zostanie wyciągnięty 2-literowy kod kraju (prefiks). EL (Grecja) → GR.",
            )
            vat_col_idx = headers.index(vat_column) + 1 if vat_column != "— nie wyciągaj —" else None

        fc3, _fc4 = st.columns(2)
        with fc3:
            cn_options = ["— nie sprawdzaj —"] + headers
            # Auto-wykryj kolumnę z kodem CN
            cn_default = 0
            for i, h in enumerate(headers):
                hl = h.strip().lower()
                if hl in ("cn", "kod cn", "kodcn", "cncode", "cn code") or "commodity" in hl or "nomenkl" in hl:
                    cn_default = i + 1
                    break
            cn_column = st.selectbox(
                f"Kolumna z kodem CN (walidacja wg CN {CN_EDITION_YEAR})", cn_options, index=cn_default,
                help=f"Każdy kod sprawdzany względem obowiązującej edycji CN {CN_EDITION_YEAR}. Kod 8-cyfrowy (CN) walidowany w całości; kod 10-cyfrowy (TARIC) — po prefiksie 8 cyfr. Status w nowej kolumnie obok: OK / nieaktualny / błędny format (z dopiskiem dla 10-cyfrowych).",
            )
            cn_col_idx = headers.index(cn_column) + 1 if cn_column != "— nie sprawdzaj —" else None
        if cn_col_idx and not VALID_CN:
            st.warning(f"Brak wbudowanej listy kodów CN {CN_EDITION_YEAR} (plik cn_{CN_EDITION_YEAR}.txt) — walidacja CN zostanie pominięta.")

        trim_multiline = st.checkbox(
            "Utnij wielolinijkowy tekst do pierwszej linii (we wszystkich komórkach)",
            value=True,
            help="Komórki ze złamaniem linii (Alt+Enter) zostaną skrócone do pierwszej linii. Jednolinijkowe pozostają bez zmian.",
        )

    # ---- Interaktywna walidacja i korekta kodów CN ----
    # Skanujemy ORYGINALNE wartości arkusza. Lista jest stabilna (kluczowana
    # oryginalnym kodem), a status "→ po poprawce" liczy się na żywo z text_input.
    cn_replacements = {}
    if cn_col_idx and VALID_CN:
        ws_scan = wb[sheet_name]
        invalid = {}  # kanoniczny kod -> {"count", "status", "sample"}
        for r in range(2, ws_scan.max_row + 1):
            raw = ws_scan.cell(row=r, column=cn_col_idx).value
            if raw is None or str(raw).strip() == "":
                continue
            code = normalize_cn(raw)
            status = cn_status(code, VALID_CN)
            if not cn_is_valid(status):
                entry = invalid.setdefault(code, {"count": 0, "status": status, "sample": raw})
                entry["count"] += 1

        if invalid:
            with st.container(border=True):
                st.markdown(f"### :red[⚠️ Niepoprawne kody CN: {len(invalid)}]")
                st.caption(
                    f"Wpisz kod zastępczy obok błędnego kodu — zastąpi go w pliku. "
                    f"Następnie kliknij „🔁 Sprawdź ponownie kody” lub od razu zapisz plik. "
                    f"Walidacja wg CN {CN_EDITION_YEAR}."
                )
                h1, h2, h3, h4 = st.columns([3, 2, 3, 2])
                h1.markdown("**Błędny kod CN**")
                h2.markdown("**Status · ile razy**")
                h3.markdown("**Kod zastępczy**")
                h4.markdown("**Po poprawce**")

                corrected = remaining = 0
                for code in sorted(invalid):
                    info = invalid[code]
                    c1, c2, c3, c4 = st.columns([3, 2, 3, 2])
                    c1.code(code, language=None)
                    c2.write(f"{info['status']} · {info['count']}×")
                    repl = c3.text_input(
                        "kod zastępczy", key=f"cn_fix_{code}",
                        label_visibility="collapsed", placeholder="np. 84713000",
                    )
                    status_now, _ = cn_outcome(code, repl, VALID_CN)
                    if repl and repl.strip():
                        cn_replacements[code] = repl.strip()
                    c4.markdown(cn_badge(status_now))
                    if status_now == "poprawiony":
                        corrected += 1
                    else:
                        remaining += 1

                b1, b2 = st.columns([1, 3])
                b1.button("🔁 Sprawdź ponownie kody")
                b2.markdown(f"**Pozostało niepoprawnych: {remaining}** · poprawionych: {corrected}")
                st.caption("Możesz teraz zapisać plik poniżej — poprawione kody zostaną podmienione w kolumnie CN.")
        else:
            st.success(f"Wszystkie kody CN w kolumnie „{cn_column}” są zgodne z CN {CN_EDITION_YEAR}.")

    # Info
    source_label = "NBP" if source == "NBP" else "EBC"
    waluta_opis = f"waluty z kolumny **\"{currency_column}\"**" if per_row_currency else f"**{currency}/PLN**"
    pln_dopisek = "; **PLN** kursem 1" if per_row_currency else ""
    if customs:
        info_text = (
            f"Dla każdego wiersza zostanie zastosowany **kurs celny NBP** {waluta_opis}: kurs ogłoszony w "
            f"**przedostatnią środę miesiąca poprzedzającego** miesiąc daty z kolumny **\"{date_column}\"** "
            f"(stały dla całego miesiąca{pln_dopisek}). Kurs trafi do nowej kolumny obok."
        )
    elif per_row_currency:
        info_text = (
            f"Dla każdego wiersza zostanie pobrany kurs {waluta_opis} względem PLN "
            f"z **ostatniego dnia roboczego przed datą** w kolumnie **\"{date_column}\"** (EUR z **{source_label}**, "
            f"pozostałe z NBP{pln_dopisek}). Kurs trafi do nowej kolumny obok."
        )
    else:
        info_text = (
            f"Kurs {waluta_opis} z **ostatniego dnia roboczego przed datą** w kolumnie **\"{date_column}\"** "
            f"zostanie pobrany z **{source_label}** i wstawiony w nowej kolumnie obok."
        )
    if amount_col_idx:
        info_text += f"\n\nKwoty z kolumny **\"{amount_column}\"** zostaną przeliczone na PLN (zaokrąglone do pełnych złotych)."
    if country_col_idx:
        info_text += f"\n\nKody krajów z kolumny **\"{country_column}\"** zostaną znormalizowane (3-literowe → 2-literowe, ISO 3166) w nowej kolumnie obok."
    if vat_col_idx:
        info_text += f"\n\nZ numerów VAT z kolumny **\"{vat_column}\"** zostanie wyciągnięty kod kraju (EL → GR) w nowej kolumnie obok."
    if cn_col_idx and VALID_CN:
        info_text += f"\n\nKody CN z kolumny **\"{cn_column}\"** zostaną sprawdzone względem obowiązującej edycji **CN {CN_EDITION_YEAR}** (status w nowej kolumnie obok)."
    if trim_multiline:
        info_text += "\n\nWielolinijkowy tekst w komórkach zostanie skrócony do pierwszej linii."
    st.info(info_text)

    # Przycisk generowania
    if st.button("Pobierz kursy i generuj plik", type="primary"):
        with st.spinner("Pobieram kursy walut..."):
            progress = st.progress(0)
            wb = process_workbook(wb, sheet_name, col_idx, source, currency, amount_col_idx, progress, country_col_idx, trim_multiline, vat_col_idx, cn_col_idx, VALID_CN, cn_replacements, currency_col_idx, rate_basis)

        st.success("Gotowe! Kursy zostały dodane.")

        original_name = uploaded_file.name.rsplit(".", 1)[0]

        if is_csv:
            # Eksport z powrotem do CSV (bez formatowania — same wartości)
            csv_bytes = workbook_to_csv_bytes(wb[sheet_name], csv_delimiter)
            st.download_button(
                label="Pobierz plik CSV z kursami",
                data=csv_bytes,
                file_name=f"{original_name}_z_kursami.csv",
                mime="text/csv",
            )
        else:
            # Usuń zewnętrzne odnośniki (powodują błędy przy otwieraniu w Excelu)
            wb._external_links = []
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            st.download_button(
                label="Pobierz plik Excel z kursami",
                data=output,
                file_name=f"{original_name}_z_kursami.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
